# -----------------------------------------------------------------------------
# DEPRECATED — superseded by parse_recipes_json.py as of 2026-04-04
# Retained for reference only. Do not run as part of the pipeline.
# -----------------------------------------------------------------------------

import os
from collections import defaultdict

import json5
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_GAME_DIR = r"C:\Games\Vintagestory"
UNPACK_DIR = r"C:\Users\Kjol\AppData\Roaming\VintagestoryData\Cache\unpack"
OUTPUT_XLSX = r"C:\Users\Kjol\Desktop\vs_recipes.xlsx"

COLUMNS = ["Source Mod", "Recipe Type", "Output Item", "Output Qty", "Ingredients"]


def parse_recipe_file(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        data = json5.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return [data]
    return []


def normalize_qty(value):
    if value is None:
        return ""
    try:
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        return str(value)
    except Exception:
        return str(value)


def read_qty(obj):
    if not isinstance(obj, dict):
        return ""
    for key in ("quantity", "stacksize", "litres", "amount"):
        if key in obj:
            return normalize_qty(obj.get(key))
    return ""


def item_code(obj):
    if isinstance(obj, str):
        return obj
    if not isinstance(obj, dict):
        return ""
    code = obj.get("code") or obj.get("name") or obj.get("type") or ""
    t = obj.get("type")
    if t and code and code != t:
        base = f"{t}:{code}"
    else:
        base = str(code)

    attributes = obj.get("attributes")
    if isinstance(attributes, dict) and attributes:
        parts = [base]
        for k in sorted(attributes.keys(), key=lambda x: str(x)):
            v = attributes.get(k)
            parts.append(f"{k}={v}")
        return "|".join(str(p) for p in parts)

    return base


def describe_stack(obj):
    if isinstance(obj, str):
        return obj
    if not isinstance(obj, dict):
        return str(obj)

    code = item_code(obj)
    qty = read_qty(obj)
    extras = []

    if "isTool" in obj and obj.get("isTool"):
        extras.append("tool")
    if "toolDurabilityCost" in obj:
        extras.append(f"dur={obj.get('toolDurabilityCost')}")
    if "minratio" in obj or "maxratio" in obj:
        extras.append(f"ratio={obj.get('minratio', '')}-{obj.get('maxratio', '')}")
    if "minQuantity" in obj or "maxQuantity" in obj:
        extras.append(f"qtyRange={obj.get('minQuantity', '')}-{obj.get('maxQuantity', '')}")
    if "allowedVariants" in obj and isinstance(obj.get("allowedVariants"), list):
        allowed = ",".join(str(v) for v in obj.get("allowedVariants", []))
        extras.append(f"variants={allowed}")

    value = f"{qty}x {code}".strip() if qty else code
    if extras:
        value = f"{value} ({'; '.join(extras)})"
    return value


def count_symbol_in_pattern(recipe, symbol):
    pattern = recipe.get("ingredientPattern") if isinstance(recipe, dict) else None
    if not isinstance(symbol, str) or not symbol:
        return 1

    if isinstance(pattern, str) and pattern:
        count = pattern.count(symbol)
        if count > 0:
            return count

    return 1


def describe_ingredients(recipe):
    parts = []

    ingredient = recipe.get("ingredient")
    if ingredient is not None:
        parts.append(describe_stack(ingredient))

    ingredients = recipe.get("ingredients")
    if isinstance(ingredients, dict):
        for k, v in ingredients.items():
            stack_desc = describe_stack(v)
            if not stack_desc:
                continue

            explicit_qty = read_qty(v) if isinstance(v, dict) else ""
            if not explicit_qty:
                slot_count = count_symbol_in_pattern(recipe, str(k))
                if slot_count > 1:
                    stack_desc = f"{slot_count}x {stack_desc}"

            parts.append(f"{k}={stack_desc}")
    elif isinstance(ingredients, list):
        for idx, ing in enumerate(ingredients, start=1):
            if isinstance(ing, dict) and isinstance(ing.get("validStacks"), list):
                valid = [describe_stack(s) for s in ing.get("validStacks", [])]
                code = ing.get("code", f"slot{idx}")
                qmin = ing.get("minQuantity")
                qmax = ing.get("maxQuantity")
                if qmin is not None or qmax is not None:
                    parts.append(f"{code}={{{' | '.join(valid)}}} ({qmin}-{qmax})")
                else:
                    parts.append(f"{code}={{{' | '.join(valid)}}}")
            else:
                parts.append(describe_stack(ing))

    if not parts:
        return ""
    return "; ".join(p for p in parts if p)


def extract_output(recipe):
    out = recipe.get("output")
    if out is None:
        out = recipe.get("cooksInto")

    if isinstance(out, list):
        if not out:
            return "", ""
        joined_codes = ", ".join(item_code(x) for x in out)
        return joined_codes, ""

    if isinstance(out, dict):
        return item_code(out), read_qty(out)

    if isinstance(out, str):
        return out, ""

    return "", ""


def infer_recipe_type(path, recipe):
    parts = [p.lower() for p in path.replace("/", "\\").split("\\")]
    if "recipes" in parts:
        i = parts.index("recipes")
        if i + 1 < len(parts):
            nxt = parts[i + 1]
            if nxt.endswith(".json"):
                nxt = "unknown"
            return nxt

    if "cooksInto" in recipe:
        return "cooking"
    if "ingredientPattern" in recipe:
        return "grid"
    if "pattern" in recipe and "ingredient" in recipe:
        return "smithing/knapping/clayforming"
    return "unknown"


def collect_files_from_base():
    candidates = []
    for rel in [r"assets\survival\recipes", r"assets\game\recipes"]:
        root = os.path.join(BASE_GAME_DIR, rel)
        if not os.path.isdir(root):
            continue
        for dirpath, _, filenames in os.walk(root):
            if "recipes" not in dirpath.lower():
                continue
            for fn in filenames:
                if fn.lower().endswith(".json"):
                    candidates.append(("Base Game", os.path.join(dirpath, fn)))
    return candidates


def collect_files_from_mods():
    candidates = []
    if not os.path.isdir(UNPACK_DIR):
        return candidates

    for mod_folder in os.listdir(UNPACK_DIR):
        mod_path = os.path.join(UNPACK_DIR, mod_folder)
        if not os.path.isdir(mod_path):
            continue
        for dirpath, _, filenames in os.walk(mod_path):
            lower = dirpath.lower()
            if "\\recipes" not in lower and "/recipes" not in lower:
                continue
            for fn in filenames:
                if fn.lower().endswith(".json"):
                    candidates.append((mod_folder, os.path.join(dirpath, fn)))
    return candidates


def collect_rows():
    rows = []
    parse_errors = []

    all_files = collect_files_from_base() + collect_files_from_mods()

    for source_mod, path in all_files:
        try:
            recipes = parse_recipe_file(path)
        except Exception as e:
            parse_errors.append((path, str(e)))
            continue

        for recipe in recipes:
            if not isinstance(recipe, dict):
                continue
            rtype = infer_recipe_type(path, recipe)
            out_item, out_qty = extract_output(recipe)
            ingredients = describe_ingredients(recipe)
            rows.append({
                "Source Mod": source_mod,
                "Recipe Type": rtype,
                "Output Item": out_item,
                "Output Qty": out_qty,
                "Ingredients": ingredients,
            })

    return rows, parse_errors


def sanitize_sheet_name(name):
    bad = set('[]:*?/\\')
    clean = "".join("_" if c in bad else c for c in str(name))
    clean = clean.strip() or "unknown"
    return clean[:31]


def autosize(ws):
    for col in range(1, len(COLUMNS) + 1):
        max_len = len(COLUMNS[col - 1])
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            val = row[0].value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 100)


def write_workbook(rows):
    wb = Workbook()
    all_ws = wb.active
    all_ws.title = "All Recipes"

    # Header
    all_ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        all_ws.cell(row=1, column=c).font = Font(bold=True)

    by_type = defaultdict(list)
    for row in rows:
        all_ws.append([row[c] for c in COLUMNS])
        by_type[row["Recipe Type"]].append(row)

    autosize(all_ws)

    for rtype in sorted(by_type.keys()):
        ws = wb.create_sheet(sanitize_sheet_name(rtype))
        ws.append(COLUMNS)
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        for row in by_type[rtype]:
            ws.append([row[c] for c in COLUMNS])
        autosize(ws)

    wb.save(OUTPUT_XLSX)
    return sorted(by_type.keys())


def main():
    rows, errors = collect_rows()
    recipe_types = write_workbook(rows)

    print(f"Wrote: {OUTPUT_XLSX}")
    print(f"Total recipes exported: {len(rows)}")
    print(f"Recipe types ({len(recipe_types)}): {', '.join(recipe_types)}")
    print(f"Parse errors skipped: {len(errors)}")

    if errors:
        print("Sample parse errors:")
        for path, err in errors[:10]:
            print(f"- {path} :: {err}")


if __name__ == "__main__":
    main()
