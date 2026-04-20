#!/usr/bin/env python3
# DEPRECATED (Gen 2 legacy): Excel-to-vs_recipes importer retained for reference only.
"""
Import Vintage Story recipes from an Excel workbook into PostgreSQL.

Default source workbook:
    C:\\Users\\Kjol\\Desktop\\vs_recipes.xlsx

Reads the "All Recipes" sheet with expected columns:
    Source Mod, Recipe Type, Output Item, Output Qty, Ingredients

Creates table + indexes, performs idempotent upsert-like insert (ON CONFLICT DO NOTHING),
and prints validation query results.
"""

from __future__ import annotations

import argparse
import getpass
import hashlib
import os
from typing import Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

try:
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError as exc:
    raise SystemExit(
        "Missing dependency 'psycopg2'. Install with: pip install psycopg2-binary"
    ) from exc


DEFAULT_XLSX = r"C:\Users\Kjol\Desktop\vs_recipes.xlsx"
DEFAULT_SHEET = "All Recipes"
TABLE_NAME = "vs_recipes"


def clean_cell(value: object) -> Optional[str]:
    """Convert blank/empty cell values to None; otherwise return stripped string."""
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def make_row_hash(
    source_mod: Optional[str],
    recipe_type: Optional[str],
    output_item: Optional[str],
    output_qty: Optional[str],
    ingredients: Optional[str],
) -> str:
    """Stable hash for idempotency, including NULL-safe normalization."""
    parts = [source_mod or "", recipe_type or "", output_item or "", output_qty or "", ingredients or ""]
    payload = "\x1f".join(parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import vs_recipes.xlsx into PostgreSQL")
    parser.add_argument("--host", default="localhost")
    parser.add_argument("--port", type=int, default=5432)
    parser.add_argument("--database", default="postgres")
    parser.add_argument("--user", default="postgres")
    parser.add_argument("--password", default=None, help="Optional. If omitted, uses PGPASSWORD or interactive prompt.")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--batch-size", type=int, default=1000)
    return parser.parse_args()


def resolve_password(cli_password: Optional[str]) -> str:
    if cli_password:
        return cli_password
    env_pw = os.getenv("PGPASSWORD")
    if env_pw:
        return env_pw
    return getpass.getpass("PostgreSQL password: ")


def load_rows_from_excel(xlsx_path: str, sheet_name: str) -> List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str], str]]:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise ValueError("Sheet is empty; no header found.")

    header_map = {str(col).strip(): idx for idx, col in enumerate(header) if col is not None}
    required = ["Source Mod", "Recipe Type", "Output Item", "Output Qty", "Ingredients"]
    missing = [c for c in required if c not in header_map]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    source_idx = header_map["Source Mod"]
    type_idx = header_map["Recipe Type"]
    output_item_idx = header_map["Output Item"]
    output_qty_idx = header_map["Output Qty"]
    ingredients_idx = header_map["Ingredients"]

    out: List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str], str]] = []

    for row in rows_iter:
        source_mod = clean_cell(row[source_idx] if source_idx < len(row) else None)
        recipe_type = clean_cell(row[type_idx] if type_idx < len(row) else None)
        output_item = clean_cell(row[output_item_idx] if output_item_idx < len(row) else None)
        output_qty = clean_cell(row[output_qty_idx] if output_qty_idx < len(row) else None)
        ingredients = clean_cell(row[ingredients_idx] if ingredients_idx < len(row) else None)

        # Keep rows even with partial data; only skip fully empty records.
        if not any([source_mod, recipe_type, output_item, output_qty, ingredients]):
            continue

        row_hash = make_row_hash(source_mod, recipe_type, output_item, output_qty, ingredients)
        out.append((source_mod, recipe_type, output_item, output_qty, ingredients, row_hash))

    return out


def batched(items: Sequence[Tuple], batch_size: int) -> Iterable[Sequence[Tuple]]:
    for i in range(0, len(items), batch_size):
        yield items[i : i + batch_size]


def ensure_schema(cur) -> None:
    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id BIGSERIAL PRIMARY KEY,
            source_mod TEXT NULL,
            recipe_type TEXT NULL,
            output_item TEXT NULL,
            output_qty TEXT NULL,
            ingredients TEXT NULL,
            row_hash CHAR(64) NOT NULL UNIQUE,
            inserted_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_recipe_type ON {TABLE_NAME} (recipe_type);")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_output_item ON {TABLE_NAME} (output_item);")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_source_mod ON {TABLE_NAME} (source_mod);")


def insert_rows(cur, rows: Sequence[Tuple], batch_size: int) -> int:
    inserted_total = 0
    sql = f"""
        INSERT INTO {TABLE_NAME} (source_mod, recipe_type, output_item, output_qty, ingredients, row_hash)
        VALUES %s
        ON CONFLICT (row_hash) DO NOTHING
    """
    for chunk in batched(rows, batch_size):
        execute_values(cur, sql, chunk, page_size=batch_size)
        inserted_total += cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
    return inserted_total


def print_validation(cur) -> None:
    print("\nValidation: total row count")
    cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME};")
    print(f"  total_rows = {cur.fetchone()[0]}")

    print("\nValidation: count by recipe type")
    cur.execute(
        f"""
        SELECT COALESCE(recipe_type, '<NULL>') AS recipe_type, COUNT(*) AS cnt
        FROM {TABLE_NAME}
        GROUP BY recipe_type
        ORDER BY cnt DESC, recipe_type ASC
        """
    )
    for recipe_type, cnt in cur.fetchall():
        print(f"  {recipe_type}: {cnt}")

    print("\nValidation: top source mods by recipe count")
    cur.execute(
        f"""
        SELECT COALESCE(source_mod, '<NULL>') AS source_mod, COUNT(*) AS cnt
        FROM {TABLE_NAME}
        GROUP BY source_mod
        ORDER BY cnt DESC, source_mod ASC
        LIMIT 15
        """
    )
    for source_mod, cnt in cur.fetchall():
        print(f"  {source_mod}: {cnt}")


def main() -> None:
    args = parse_args()
    password = resolve_password(args.password)

    rows = load_rows_from_excel(args.xlsx, args.sheet)
    print(f"Loaded {len(rows)} rows from '{args.xlsx}' [{args.sheet}].")

    conn = psycopg2.connect(
        host=args.host,
        port=args.port,
        dbname=args.database,
        user=args.user,
        password=password,
    )

    try:
        with conn:
            with conn.cursor() as cur:
                ensure_schema(cur)
                inserted = insert_rows(cur, rows, args.batch_size)
                print(f"Inserted {inserted} new rows (duplicates skipped via row_hash).")
                print_validation(cur)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
