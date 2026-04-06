#!/usr/bin/env python3
"""
Ingest Faelyn Trade Association (FTA) price sheets into fta_items.

Reads DATABASE_URL from environment and parses the local workbook:
  C:/Users/Kjol/Desktop/Vintage Story Modded Multiplayer/Calculator Data/FTA Spreadsheets/Copy of Pie baron FTA price sheet.xlsx

Uses only category sheets:
  - Agriculture
  - Food
  - Industrial
  - Artisan
"""

import csv
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from typing import Dict, Iterable, List, Optional, Tuple

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()


WORKBOOK_PATH = (
    r"C:\Users\Kjol\Desktop\Vintage Story Modded Multiplayer\Calculator Data"
    r"\FTA Spreadsheets\Copy of Pie baron FTA price sheet.xlsx"
)

SHEETS: List[Tuple[str, str]] = [
    ("Agriculture", "agriculture"),
    ("Food", "food"),
    ("Industrial", "industrial"),
    ("Artisan", "artisan"),
]

ITEM_ID_RE = re.compile(r"^[A-Z]{2}\d{3}$")
LITER_RE = re.compile(r"(?i)\b(\d+(?:\.\d+)?)\s*L\b")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def decimal_from_value(value: object) -> Optional[Decimal]:
    text = normalize_text(value)
    if text == "":
        return None

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    cleaned = text.replace(",", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def normalize_qty_unit(raw_value: object) -> str:
    if raw_value is None:
        return ""

    if isinstance(raw_value, float) and raw_value.is_integer():
        return str(int(raw_value))

    text = str(raw_value).strip()
    if text.endswith(".0"):
        maybe_num = decimal_from_value(text)
        if maybe_num is not None and maybe_num == maybe_num.to_integral_value():
            return str(int(maybe_num))
    return text


def parse_qty_fields(qty_unit: str) -> Tuple[Optional[Decimal], bool]:
    text = qty_unit.strip()
    if text == "":
        return None, False

    liter_match = LITER_RE.search(text)
    if liter_match:
        return Decimal(liter_match.group(1)), True

    qty_numeric = decimal_from_value(text)
    return qty_numeric, False


def parse_price_fields(price_value_raw: object) -> Tuple[Optional[Decimal], Optional[str], bool]:
    """
    Returns:
      (copper_sovereign_value, notes, should_skip)
    """
    text = normalize_text(price_value_raw)
    if text == "" or text == "-":
        return None, None, True

    numeric_value = decimal_from_value(price_value_raw)
    if numeric_value is not None:
        if numeric_value == 0:
            return None, None, True
        return numeric_value, None, False

    # Non-numeric text prices (e.g. "1/4 ingot price") are preserved in notes.
    return None, text, False


def save_debug_csv(sheet_name: str, debug_rows: List[Dict[str, object]]) -> str:
    raw_dir = os.path.join("data", "raw")
    os.makedirs(raw_dir, exist_ok=True)

    slug = re.sub(r"[^a-z0-9]+", "_", sheet_name.lower()).strip("_")
    output_path = os.path.join(raw_dir, f"fta_{slug}.csv")

    fieldnames = [
        "row_number",
        "item_id_raw",
        "name_raw",
        "qty_unit_raw",
        "copper_value_raw",
        "fta_sub_category",
        "action",
        "skip_reason",
        "item_id",
        "display_name",
        "qty_unit",
        "qty_numeric",
        "unit_is_liters",
        "copper_sovereign_value",
        "notes",
    ]

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in debug_rows:
            writer.writerow(row)

    return output_path


def parse_sheet(ws, fta_category: str) -> Tuple[List[Dict[str, object]], Dict[str, int], List[Dict[str, object]]]:
    parsed_rows: List[Dict[str, object]] = []
    debug_rows: List[Dict[str, object]] = []

    stats = {
        "rows_scanned": 0,
        "non_data_rows": 0,
        "parsed": 0,
        "skipped": 0,
    }

    current_sub_category: Optional[str] = None

    for row_number, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        stats["rows_scanned"] += 1

        item_id_raw = row[0] if len(row) >= 1 else None
        name_raw = row[1] if len(row) >= 2 else None
        qty_unit_raw = row[2] if len(row) >= 3 else None
        copper_raw = row[3] if len(row) >= 4 else None

        item_id = normalize_text(item_id_raw).upper()
        name_text = normalize_text(name_raw)

        if not ITEM_ID_RE.match(item_id):
            # Section headers are non-item rows with text in column B.
            if name_text and "ITEM ID" not in name_text.upper():
                current_sub_category = name_text.strip(" ,") or current_sub_category

            stats["non_data_rows"] += 1
            debug_rows.append(
                {
                    "row_number": row_number,
                    "item_id_raw": normalize_text(item_id_raw),
                    "name_raw": name_text,
                    "qty_unit_raw": normalize_text(qty_unit_raw),
                    "copper_value_raw": normalize_text(copper_raw),
                    "fta_sub_category": current_sub_category or "",
                    "action": "skip",
                    "skip_reason": "non_data_row",
                    "item_id": "",
                    "display_name": "",
                    "qty_unit": "",
                    "qty_numeric": "",
                    "unit_is_liters": "",
                    "copper_sovereign_value": "",
                    "notes": "",
                }
            )
            continue

        qty_unit = normalize_qty_unit(qty_unit_raw)
        qty_numeric, unit_is_liters = parse_qty_fields(qty_unit)
        copper_value, notes, should_skip = parse_price_fields(copper_raw)

        if should_skip:
            stats["skipped"] += 1
            debug_rows.append(
                {
                    "row_number": row_number,
                    "item_id_raw": normalize_text(item_id_raw),
                    "name_raw": name_text,
                    "qty_unit_raw": normalize_text(qty_unit_raw),
                    "copper_value_raw": normalize_text(copper_raw),
                    "fta_sub_category": current_sub_category or "",
                    "action": "skip",
                    "skip_reason": "blank_dash_or_zero_price",
                    "item_id": item_id,
                    "display_name": name_text,
                    "qty_unit": qty_unit,
                    "qty_numeric": str(qty_numeric) if qty_numeric is not None else "",
                    "unit_is_liters": str(unit_is_liters),
                    "copper_sovereign_value": "",
                    "notes": notes or "",
                }
            )
            continue

        row_dict = {
            "item_id": item_id,
            "display_name": name_text,
            "fta_category": fta_category,
            "fta_sub_category": current_sub_category,
            "qty_unit": qty_unit,
            "qty_numeric": qty_numeric,
            "unit_is_liters": unit_is_liters,
            "copper_sovereign_value": copper_value,
            "notes": notes,
        }
        parsed_rows.append(row_dict)
        stats["parsed"] += 1

        debug_rows.append(
            {
                "row_number": row_number,
                "item_id_raw": normalize_text(item_id_raw),
                "name_raw": name_text,
                "qty_unit_raw": normalize_text(qty_unit_raw),
                "copper_value_raw": normalize_text(copper_raw),
                "fta_sub_category": current_sub_category or "",
                "action": "parse",
                "skip_reason": "",
                "item_id": item_id,
                "display_name": name_text,
                "qty_unit": qty_unit,
                "qty_numeric": str(qty_numeric) if qty_numeric is not None else "",
                "unit_is_liters": str(unit_is_liters),
                "copper_sovereign_value": str(copper_value) if copper_value is not None else "",
                "notes": notes or "",
            }
        )

    return parsed_rows, stats, debug_rows


UPSERT_SQL = """
INSERT INTO fta_items (
    item_id,
    display_name,
    fta_category,
    fta_sub_category,
    qty_unit,
    qty_numeric,
    unit_is_liters,
    copper_sovereign_value,
    notes
) VALUES (
    %(item_id)s,
    %(display_name)s,
    %(fta_category)s,
    %(fta_sub_category)s,
    %(qty_unit)s,
    %(qty_numeric)s,
    %(unit_is_liters)s,
    %(copper_sovereign_value)s,
    %(notes)s
)
ON CONFLICT (item_id) DO UPDATE
SET
    display_name = EXCLUDED.display_name,
    fta_category = EXCLUDED.fta_category,
    fta_sub_category = EXCLUDED.fta_sub_category,
    qty_unit = EXCLUDED.qty_unit,
    qty_numeric = EXCLUDED.qty_numeric,
    unit_is_liters = EXCLUDED.unit_is_liters,
    copper_sovereign_value = EXCLUDED.copper_sovereign_value,
    notes = EXCLUDED.notes
RETURNING (xmax = 0) AS inserted;
"""


def ingest_rows(cur, rows: Iterable[Dict[str, object]]) -> Tuple[int, int]:
    inserted = 0
    updated = 0

    for row in rows:
        cur.execute(UPSERT_SQL, row)
        was_inserted = cur.fetchone()[0]
        if was_inserted:
            inserted += 1
        else:
            updated += 1

    return inserted, updated


def main() -> int:
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        print("[ERROR] DATABASE_URL environment variable is not set.", file=sys.stderr)
        return 1

    if not os.path.exists(WORKBOOK_PATH):
        print(f"[ERROR] Workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    print("Starting FTA price ingestion...")
    print(f"Workbook: {WORKBOOK_PATH}")

    try:
        wb = load_workbook(WORKBOOK_PATH, data_only=False)
    except Exception as exc:
        print(f"[ERROR] Failed to open workbook: {exc}", file=sys.stderr)
        return 1

    sheet_rows: Dict[str, List[Dict[str, object]]] = {}
    sheet_stats: Dict[str, Dict[str, int]] = {}

    for sheet_name, category_slug in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet not found (skipping): {sheet_name}", file=sys.stderr)
            continue

        print(f"\nParsing sheet: {sheet_name}")
        ws = wb[sheet_name]
        parsed_rows, stats, debug_rows = parse_sheet(ws, category_slug)
        debug_path = save_debug_csv(sheet_name, debug_rows)

        sheet_rows[sheet_name] = parsed_rows
        sheet_stats[sheet_name] = stats

        print(f"  Saved debug CSV: {debug_path}")
        print(
            f"  Parsed={stats['parsed']} Skipped={stats['skipped']} "
            f"(Non-data rows={stats['non_data_rows']})"
        )

    if not sheet_rows:
        print("[ERROR] No sheets were parsed; nothing to ingest.", file=sys.stderr)
        return 1

    try:
        with psycopg2.connect(database_url) as conn:
            with conn.cursor() as cur:
                db_totals: Dict[str, Tuple[int, int]] = {}
                for sheet_name, rows in sheet_rows.items():
                    ins, upd = ingest_rows(cur, rows)
                    db_totals[sheet_name] = (ins, upd)

                print("\nIngestion summary:")
                for sheet_name, stats in sheet_stats.items():
                    ins, upd = db_totals.get(sheet_name, (0, 0))
                    print(
                        f"  {sheet_name}: parsed={stats['parsed']}, skipped={stats['skipped']}, "
                        f"inserted={ins}, updated={upd}"
                    )

    except Exception as exc:
        print(f"[ERROR] Database ingestion failed: {exc}", file=sys.stderr)
        return 1

    print("\nDone.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
