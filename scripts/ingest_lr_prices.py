#!/usr/bin/env python3
"""
Ingest Lost Realm economy price workbook into lr_items.

Reads DATABASE_URL from environment and parses local workbook:
  data/raw/lr_economy_workbook.xlsx
"""

import csv
import os
import re
import sys
from collections import Counter
from decimal import Decimal, InvalidOperation
from typing import Dict, Iterable, List, Optional, Tuple

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()


WORKBOOK_PATH = os.path.join("data", "raw", "lr_economy_workbook.xlsx")

SHEETS: List[Tuple[str, str]] = [
    ("AGRICULTURAL GOODS", "agricultural_goods"),
    ("INDUSTRIAL GOODS", "industrial_goods"),
    ("ARTISANAL GOODS", "artisanal_goods"),
    ("Settlement Specialization", "settlement_specialization"),
]

ITEM_ID_RE = re.compile(r"^[A-Z]{2,}\d+$")

STANDARD_COL_MAPS: Dict[str, Dict[str, int]] = {
    # Item ID, Name, Count, Base Value, Last Value, Current Price, settlements...
    "agricultural_goods": {
        "current_price": 5,
        "industrial_town": 6,
        "industrial_city": 7,
        "market_town": 8,
        "market_city": 9,
        "religious_town": 10,
        "temple_city": 11,
    },
    # Item ID, Name, Count, Base Value, NA, Last Value, Current Price, settlements...
    "industrial_goods": {
        "current_price": 6,
        "industrial_town": 7,
        "industrial_city": 8,
        "market_town": 9,
        "market_city": 10,
        "religious_town": 11,
        "temple_city": 12,
    },
}

STANDARD_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "current_price": ("current price",),
    "industrial_town": ("industrial town",),
    "industrial_city": ("industrial city",),
    "market_town": ("market town",),
    "market_city": ("market city",),
    "religious_town": ("religious town",),
    "temple_city": ("temple city",),
}


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_numeric(value: object) -> Optional[Decimal]:
    text = normalize_cell(value)
    if text in ("", "-"):
        return None

    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid numeric value: {value!r}") from exc


def parse_numeric_loose(value: object) -> Optional[Decimal]:
    try:
        return parse_numeric(value)
    except ValueError:
        return None


def parse_int(value: object) -> int:
    text = normalize_cell(value)
    if text == "":
        raise ValueError("Count is blank")

    text = text.replace(",", "")
    match = re.match(r"^(\d+)", text)
    if not match:
        raise ValueError(f"Invalid integer value for count: {value!r}")

    return int(match.group(1))


def get_col(row: List[object], idx: int) -> object:
    if idx < len(row):
        return row[idx]
    return None


def resolve_standard_col_map(ws, category: str) -> Optional[Dict[str, int]]:
    """Resolve standard category column indexes from header labels.

    This ensures we use the workbook's *Current Price* column by name rather
    than relying only on fixed index assumptions.
    """
    default_map = STANDARD_COL_MAPS.get(category)
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return default_map

    header_lookup: Dict[str, int] = {}
    for idx, value in enumerate(header_row):
        label = normalize_cell(value).lower()
        if label and label not in header_lookup:
            header_lookup[label] = idx

    resolved: Dict[str, int] = {}
    for logical_key, aliases in STANDARD_HEADER_ALIASES.items():
        found_idx = None
        for alias in aliases:
            if alias in header_lookup:
                found_idx = header_lookup[alias]
                break
        if found_idx is None:
            if default_map is None:
                return None
            found_idx = default_map[logical_key]
        resolved[logical_key] = found_idx

    return resolved


def extract_sub_category(row: List[object], category: str) -> Optional[str]:
    second_col = normalize_cell(get_col(row, 1))
    if not second_col:
        return None

    upper = second_col.upper()
    if "ITEM ID" in upper:
        return None
    if upper in {
        "INDUSTRIAL GOODS",
        "AGRICULTURAL GOODS",
        "ARTISANAL GOODS",
        "SETTLEMENT SPECIALIZATION",
    }:
        return None
    if second_col.startswith("Notes:"):
        return None
    if category == "settlement_specialization" and second_col in {
        "Category Multiplied",
        "Agricultural  Multiplier",
        "Industrial Multiplier",
        "Artisanal Multiplier",
    }:
        return None

    return second_col.strip(" ,") or None


def parse_standard_row(
    row: List[object],
    category: str,
    sub_category: Optional[str],
    standard_col_map: Dict[str, int],
) -> Dict[str, object]:
    has_quality_tiers = any(parse_numeric_loose(get_col(row, idx)) is not None for idx in (17, 28, 39))

    if has_quality_tiers:
        return {
            "item_id": normalize_cell(get_col(row, 0)).upper(),
            "display_name": normalize_cell(get_col(row, 1)),
            "lr_category": category,
            "lr_sub_category": sub_category,
            "count": parse_int(get_col(row, 2)),
            "base_value": parse_numeric_loose(get_col(row, 3)),
            "price_current": parse_numeric_loose(get_col(row, 6)),
            "price_industrial_town": parse_numeric_loose(get_col(row, 7)),
            "price_industrial_city": parse_numeric_loose(get_col(row, 8)),
            "price_market_town": parse_numeric_loose(get_col(row, 9)),
            "price_market_city": parse_numeric_loose(get_col(row, 10)),
            "price_religious_town": parse_numeric_loose(get_col(row, 11)),
            "price_temple_city": parse_numeric_loose(get_col(row, 12)),
            "has_quality_tiers": True,
            "price_uncommon_current": parse_numeric_loose(get_col(row, 6)),
            "price_uncommon_industrial_town": parse_numeric_loose(get_col(row, 7)),
            "price_uncommon_industrial_city": parse_numeric_loose(get_col(row, 8)),
            "price_uncommon_market_town": parse_numeric_loose(get_col(row, 9)),
            "price_uncommon_market_city": parse_numeric_loose(get_col(row, 10)),
            "price_uncommon_religious_town": parse_numeric_loose(get_col(row, 11)),
            "price_uncommon_temple_city": parse_numeric_loose(get_col(row, 12)),
            "price_rare_current": parse_numeric_loose(get_col(row, 17)),
            "price_rare_industrial_town": parse_numeric_loose(get_col(row, 18)),
            "price_rare_industrial_city": parse_numeric_loose(get_col(row, 19)),
            "price_rare_market_town": parse_numeric_loose(get_col(row, 20)),
            "price_rare_market_city": parse_numeric_loose(get_col(row, 21)),
            "price_rare_religious_town": parse_numeric_loose(get_col(row, 22)),
            "price_rare_temple_city": parse_numeric_loose(get_col(row, 23)),
            "price_epic_current": parse_numeric_loose(get_col(row, 28)),
            "price_epic_industrial_town": parse_numeric_loose(get_col(row, 29)),
            "price_epic_industrial_city": parse_numeric_loose(get_col(row, 30)),
            "price_epic_market_town": parse_numeric_loose(get_col(row, 31)),
            "price_epic_market_city": parse_numeric_loose(get_col(row, 32)),
            "price_epic_religious_town": parse_numeric_loose(get_col(row, 33)),
            "price_epic_temple_city": parse_numeric_loose(get_col(row, 34)),
            "price_legendary_current": parse_numeric_loose(get_col(row, 39)),
            "price_legendary_industrial_town": parse_numeric_loose(get_col(row, 40)),
            "price_legendary_industrial_city": parse_numeric_loose(get_col(row, 41)),
            "price_legendary_market_town": parse_numeric_loose(get_col(row, 42)),
            "price_legendary_market_city": parse_numeric_loose(get_col(row, 43)),
            "price_legendary_religious_town": parse_numeric_loose(get_col(row, 44)),
            "price_legendary_temple_city": parse_numeric_loose(get_col(row, 45)),
        }

    return {
        "item_id": normalize_cell(get_col(row, 0)).upper(),
        "display_name": normalize_cell(get_col(row, 1)),
        "lr_category": category,
        "lr_sub_category": sub_category,
        "count": parse_int(get_col(row, 2)),
        "base_value": parse_numeric_loose(get_col(row, 3)),
        "price_current": parse_numeric_loose(get_col(row, standard_col_map["current_price"])),
        "price_industrial_town": parse_numeric_loose(get_col(row, standard_col_map["industrial_town"])),
        "price_industrial_city": parse_numeric_loose(get_col(row, standard_col_map["industrial_city"])),
        "price_market_town": parse_numeric_loose(get_col(row, standard_col_map["market_town"])),
        "price_market_city": parse_numeric_loose(get_col(row, standard_col_map["market_city"])),
        "price_religious_town": parse_numeric_loose(get_col(row, standard_col_map["religious_town"])),
        "price_temple_city": parse_numeric_loose(get_col(row, standard_col_map["temple_city"])),
        "has_quality_tiers": False,
        "price_uncommon_current": None,
        "price_uncommon_industrial_town": None,
        "price_uncommon_industrial_city": None,
        "price_uncommon_market_town": None,
        "price_uncommon_market_city": None,
        "price_uncommon_religious_town": None,
        "price_uncommon_temple_city": None,
        "price_rare_current": None,
        "price_rare_industrial_town": None,
        "price_rare_industrial_city": None,
        "price_rare_market_town": None,
        "price_rare_market_city": None,
        "price_rare_religious_town": None,
        "price_rare_temple_city": None,
        "price_epic_current": None,
        "price_epic_industrial_town": None,
        "price_epic_industrial_city": None,
        "price_epic_market_town": None,
        "price_epic_market_city": None,
        "price_epic_religious_town": None,
        "price_epic_temple_city": None,
        "price_legendary_current": None,
        "price_legendary_industrial_town": None,
        "price_legendary_industrial_city": None,
        "price_legendary_market_town": None,
        "price_legendary_market_city": None,
        "price_legendary_religious_town": None,
        "price_legendary_temple_city": None,
    }


def parse_artisanal_row(row: List[object], category: str, sub_category: Optional[str]) -> Dict[str, object]:
    item_id = normalize_cell(get_col(row, 0)).upper()
    count = parse_int(get_col(row, 2))
    base_value = parse_numeric_loose(get_col(row, 3))

    has_quality_tiers = any(parse_numeric_loose(get_col(row, idx)) is not None for idx in (17, 28, 39))

    if has_quality_tiers:
        return {
            "item_id": item_id,
            "display_name": normalize_cell(get_col(row, 1)),
            "lr_category": category,
            "lr_sub_category": sub_category,
            "count": count,
            "base_value": base_value,
            "price_current": parse_numeric_loose(get_col(row, 6)),
            "price_industrial_town": parse_numeric_loose(get_col(row, 7)),
            "price_industrial_city": parse_numeric_loose(get_col(row, 8)),
            "price_market_town": parse_numeric_loose(get_col(row, 9)),
            "price_market_city": parse_numeric_loose(get_col(row, 10)),
            "price_religious_town": parse_numeric_loose(get_col(row, 11)),
            "price_temple_city": parse_numeric_loose(get_col(row, 12)),
            "has_quality_tiers": True,
            "price_uncommon_current": parse_numeric_loose(get_col(row, 6)),
            "price_uncommon_industrial_town": parse_numeric_loose(get_col(row, 7)),
            "price_uncommon_industrial_city": parse_numeric_loose(get_col(row, 8)),
            "price_uncommon_market_town": parse_numeric_loose(get_col(row, 9)),
            "price_uncommon_market_city": parse_numeric_loose(get_col(row, 10)),
            "price_uncommon_religious_town": parse_numeric_loose(get_col(row, 11)),
            "price_uncommon_temple_city": parse_numeric_loose(get_col(row, 12)),
            "price_rare_current": parse_numeric_loose(get_col(row, 17)),
            "price_rare_industrial_town": parse_numeric_loose(get_col(row, 18)),
            "price_rare_industrial_city": parse_numeric_loose(get_col(row, 19)),
            "price_rare_market_town": parse_numeric_loose(get_col(row, 20)),
            "price_rare_market_city": parse_numeric_loose(get_col(row, 21)),
            "price_rare_religious_town": parse_numeric_loose(get_col(row, 22)),
            "price_rare_temple_city": parse_numeric_loose(get_col(row, 23)),
            "price_epic_current": parse_numeric_loose(get_col(row, 28)),
            "price_epic_industrial_town": parse_numeric_loose(get_col(row, 29)),
            "price_epic_industrial_city": parse_numeric_loose(get_col(row, 30)),
            "price_epic_market_town": parse_numeric_loose(get_col(row, 31)),
            "price_epic_market_city": parse_numeric_loose(get_col(row, 32)),
            "price_epic_religious_town": parse_numeric_loose(get_col(row, 33)),
            "price_epic_temple_city": parse_numeric_loose(get_col(row, 34)),
            "price_legendary_current": parse_numeric_loose(get_col(row, 39)),
            "price_legendary_industrial_town": parse_numeric_loose(get_col(row, 40)),
            "price_legendary_industrial_city": parse_numeric_loose(get_col(row, 41)),
            "price_legendary_market_town": parse_numeric_loose(get_col(row, 42)),
            "price_legendary_market_city": parse_numeric_loose(get_col(row, 43)),
            "price_legendary_religious_town": parse_numeric_loose(get_col(row, 44)),
            "price_legendary_temple_city": parse_numeric_loose(get_col(row, 45)),
        }

    return {
        "item_id": item_id,
        "display_name": normalize_cell(get_col(row, 1)),
        "lr_category": category,
        "lr_sub_category": sub_category,
        "count": count,
        "base_value": base_value,
        "price_current": parse_numeric_loose(get_col(row, 6)),
        "price_industrial_town": parse_numeric_loose(get_col(row, 7)),
        "price_industrial_city": parse_numeric_loose(get_col(row, 8)),
        "price_market_town": parse_numeric_loose(get_col(row, 9)),
        "price_market_city": parse_numeric_loose(get_col(row, 10)),
        "price_religious_town": parse_numeric_loose(get_col(row, 11)),
        "price_temple_city": parse_numeric_loose(get_col(row, 12)),
        "has_quality_tiers": False,
        "price_uncommon_current": None,
        "price_uncommon_industrial_town": None,
        "price_uncommon_industrial_city": None,
        "price_uncommon_market_town": None,
        "price_uncommon_market_city": None,
        "price_uncommon_religious_town": None,
        "price_uncommon_temple_city": None,
        "price_rare_current": None,
        "price_rare_industrial_town": None,
        "price_rare_industrial_city": None,
        "price_rare_market_town": None,
        "price_rare_market_city": None,
        "price_rare_religious_town": None,
        "price_rare_temple_city": None,
        "price_epic_current": None,
        "price_epic_industrial_town": None,
        "price_epic_industrial_city": None,
        "price_epic_market_town": None,
        "price_epic_market_city": None,
        "price_epic_religious_town": None,
        "price_epic_temple_city": None,
        "price_legendary_current": None,
        "price_legendary_industrial_town": None,
        "price_legendary_industrial_city": None,
        "price_legendary_market_town": None,
        "price_legendary_market_city": None,
        "price_legendary_religious_town": None,
        "price_legendary_temple_city": None,
    }


def save_debug_csv(category: str, rows: List[Dict[str, object]]) -> str:
    raw_dir = os.path.join("data", "raw", "deprecated_prices_20260416")
    os.makedirs(raw_dir, exist_ok=True)
    output_path = os.path.join(raw_dir, f"lr_{category}_parsed.csv")

    if not rows:
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["note"])
            writer.writerow(["no parsed rows"])
        return output_path

    fieldnames = list(rows[0].keys())
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return output_path


def dedupe_rows_by_item_id(rows: List[Dict[str, object]], category: str) -> Tuple[List[Dict[str, object]], int]:
    """Keep first occurrence for duplicate item IDs and warn.

    Some workbook revisions contain duplicate item IDs across sections (e.g. IN067,
    IN068 in industrial). We keep the first row encountered to preserve earlier
    canonical sections and avoid silent overwrite order dependence during UPSERT.
    """
    seen = set()
    deduped: List[Dict[str, object]] = []
    duplicates = 0

    counts = Counter(str(r.get("item_id", "")) for r in rows)
    dup_ids = [item_id for item_id, count in counts.items() if item_id and count > 1]
    if dup_ids:
        print(
            f"  [WARN] {category}: duplicate item_id values detected: {dup_ids}. Keeping first occurrence.",
            file=sys.stderr,
        )

    for row in rows:
        item_id = str(row.get("item_id", ""))
        if item_id in seen:
            duplicates += 1
            continue
        seen.add(item_id)
        deduped.append(row)

    return deduped, duplicates


def parse_sheet(ws, category: str) -> Tuple[List[Dict[str, object]], int]:
    if category == "settlement_specialization":
        # Settlement multipliers are handled by a dedicated script; retain sheet parse no-op.
        scanned = sum(1 for _ in ws.iter_rows(min_row=1, values_only=True))
        return [], scanned

    standard_col_map = resolve_standard_col_map(ws, category)
    parsed: List[Dict[str, object]] = []
    skipped = 0
    current_sub_category: Optional[str] = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        row_vals: List[object] = list(row)
        item_id_cell = normalize_cell(get_col(row_vals, 0)).upper()

        if not ITEM_ID_RE.match(item_id_cell):
            maybe_sub_category = extract_sub_category(row_vals, category)
            if maybe_sub_category:
                current_sub_category = maybe_sub_category
            skipped += 1
            continue

        try:
            if category == "artisanal_goods":
                parsed.append(parse_artisanal_row(row_vals, category, current_sub_category))
            else:
                if standard_col_map is None:
                    skipped += 1
                    continue
                parsed.append(parse_standard_row(row_vals, category, current_sub_category, standard_col_map))
        except ValueError as err:
            skipped += 1
            print(f"  [WARN] Skipping malformed row ({category}): {err}", file=sys.stderr)

    return parsed, skipped


UPSERT_SQL = """
INSERT INTO lr_items (
    item_id,
    display_name,
    lr_category,
    lr_sub_category,
    count,
    base_value,
    price_current,
    price_industrial_town,
    price_industrial_city,
    price_market_town,
    price_market_city,
    price_religious_town,
    price_temple_city,
    has_quality_tiers,
    price_uncommon_current,
    price_uncommon_industrial_town,
    price_uncommon_industrial_city,
    price_uncommon_market_town,
    price_uncommon_market_city,
    price_uncommon_religious_town,
    price_uncommon_temple_city,
    price_rare_current,
    price_rare_industrial_town,
    price_rare_industrial_city,
    price_rare_market_town,
    price_rare_market_city,
    price_rare_religious_town,
    price_rare_temple_city,
    price_epic_current,
    price_epic_industrial_town,
    price_epic_industrial_city,
    price_epic_market_town,
    price_epic_market_city,
    price_epic_religious_town,
    price_epic_temple_city,
    price_legendary_current,
    price_legendary_industrial_town,
    price_legendary_industrial_city,
    price_legendary_market_town,
    price_legendary_market_city,
    price_legendary_religious_town,
    price_legendary_temple_city
) VALUES (
    %(item_id)s,
    %(display_name)s,
    %(lr_category)s,
    %(lr_sub_category)s,
    %(count)s,
    %(base_value)s,
    %(price_current)s,
    %(price_industrial_town)s,
    %(price_industrial_city)s,
    %(price_market_town)s,
    %(price_market_city)s,
    %(price_religious_town)s,
    %(price_temple_city)s,
    %(has_quality_tiers)s,
    %(price_uncommon_current)s,
    %(price_uncommon_industrial_town)s,
    %(price_uncommon_industrial_city)s,
    %(price_uncommon_market_town)s,
    %(price_uncommon_market_city)s,
    %(price_uncommon_religious_town)s,
    %(price_uncommon_temple_city)s,
    %(price_rare_current)s,
    %(price_rare_industrial_town)s,
    %(price_rare_industrial_city)s,
    %(price_rare_market_town)s,
    %(price_rare_market_city)s,
    %(price_rare_religious_town)s,
    %(price_rare_temple_city)s,
    %(price_epic_current)s,
    %(price_epic_industrial_town)s,
    %(price_epic_industrial_city)s,
    %(price_epic_market_town)s,
    %(price_epic_market_city)s,
    %(price_epic_religious_town)s,
    %(price_epic_temple_city)s,
    %(price_legendary_current)s,
    %(price_legendary_industrial_town)s,
    %(price_legendary_industrial_city)s,
    %(price_legendary_market_town)s,
    %(price_legendary_market_city)s,
    %(price_legendary_religious_town)s,
    %(price_legendary_temple_city)s
)
ON CONFLICT (item_id) DO UPDATE
SET
    display_name = EXCLUDED.display_name,
    lr_category = EXCLUDED.lr_category,
    lr_sub_category = EXCLUDED.lr_sub_category,
    count = EXCLUDED.count,
    base_value = EXCLUDED.base_value,
    price_current = EXCLUDED.price_current,
    price_industrial_town = EXCLUDED.price_industrial_town,
    price_industrial_city = EXCLUDED.price_industrial_city,
    price_market_town = EXCLUDED.price_market_town,
    price_market_city = EXCLUDED.price_market_city,
    price_religious_town = EXCLUDED.price_religious_town,
    price_temple_city = EXCLUDED.price_temple_city,
    has_quality_tiers = EXCLUDED.has_quality_tiers,
    price_uncommon_current = EXCLUDED.price_uncommon_current,
    price_uncommon_industrial_town = EXCLUDED.price_uncommon_industrial_town,
    price_uncommon_industrial_city = EXCLUDED.price_uncommon_industrial_city,
    price_uncommon_market_town = EXCLUDED.price_uncommon_market_town,
    price_uncommon_market_city = EXCLUDED.price_uncommon_market_city,
    price_uncommon_religious_town = EXCLUDED.price_uncommon_religious_town,
    price_uncommon_temple_city = EXCLUDED.price_uncommon_temple_city,
    price_rare_current = EXCLUDED.price_rare_current,
    price_rare_industrial_town = EXCLUDED.price_rare_industrial_town,
    price_rare_industrial_city = EXCLUDED.price_rare_industrial_city,
    price_rare_market_town = EXCLUDED.price_rare_market_town,
    price_rare_market_city = EXCLUDED.price_rare_market_city,
    price_rare_religious_town = EXCLUDED.price_rare_religious_town,
    price_rare_temple_city = EXCLUDED.price_rare_temple_city,
    price_epic_current = EXCLUDED.price_epic_current,
    price_epic_industrial_town = EXCLUDED.price_epic_industrial_town,
    price_epic_industrial_city = EXCLUDED.price_epic_industrial_city,
    price_epic_market_town = EXCLUDED.price_epic_market_town,
    price_epic_market_city = EXCLUDED.price_epic_market_city,
    price_epic_religious_town = EXCLUDED.price_epic_religious_town,
    price_epic_temple_city = EXCLUDED.price_epic_temple_city,
    price_legendary_current = EXCLUDED.price_legendary_current,
    price_legendary_industrial_town = EXCLUDED.price_legendary_industrial_town,
    price_legendary_industrial_city = EXCLUDED.price_legendary_industrial_city,
    price_legendary_market_town = EXCLUDED.price_legendary_market_town,
    price_legendary_market_city = EXCLUDED.price_legendary_market_city,
    price_legendary_religious_town = EXCLUDED.price_legendary_religious_town,
    price_legendary_temple_city = EXCLUDED.price_legendary_temple_city
RETURNING (xmax = 0) AS inserted;
"""


def ingest_category(cur, rows: Iterable[Dict[str, object]]) -> Tuple[int, int]:
    inserted = 0
    updated = 0
    for row in rows:
        cur.execute(UPSERT_SQL, row)
        was_inserted = cur.fetchone()[0]
        if was_inserted:
            inserted += 1
        else:
            updated += 1
    return inserted, updated


def main() -> int:
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        print("[ERROR] DATABASE_URL environment variable is not set.", file=sys.stderr)
        return 1

    if not os.path.exists(WORKBOOK_PATH):
        print(f"[ERROR] Workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    print("Starting Lost Realm price ingestion...")
    print(f"Workbook: {WORKBOOK_PATH}")

    try:
        wb = load_workbook(WORKBOOK_PATH, data_only=True)
    except Exception as exc:
        print(f"[ERROR] Failed to open workbook: {exc}", file=sys.stderr)
        return 1

    all_rows: Dict[str, List[Dict[str, object]]] = {}
    sheet_stats: Dict[str, Tuple[int, int]] = {}

    for sheet_name, category in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet not found (skipping): {sheet_name}", file=sys.stderr)
            continue

        print(f"\nParsing sheet: {sheet_name}")
        ws = wb[sheet_name]
        parsed_rows, skipped = parse_sheet(ws, category)
        deduped_rows, deduped_count = dedupe_rows_by_item_id(parsed_rows, category)
        debug_path = save_debug_csv(category, parsed_rows)
        all_rows[category] = deduped_rows
        sheet_stats[category] = (len(parsed_rows), skipped + deduped_count)
        print(f"  Saved debug CSV: {debug_path}")
        print(
            f"  Parsed {len(deduped_rows)} rows"
            + (
                f" (skipped {skipped + deduped_count})"
                if (skipped + deduped_count)
                else ""
            )
        )

    data_categories = {k: v for k, v in all_rows.items() if k != "settlement_specialization"}
    if not data_categories:
        print("[ERROR] No LR item categories were parsed; nothing to ingest.", file=sys.stderr)
        return 1

    try:
        with psycopg2.connect(database_url) as conn:
            with conn.cursor() as cur:
                totals: Dict[str, Tuple[int, int]] = {}
                for category, rows in data_categories.items():
                    ins, upd = ingest_category(cur, rows)
                    totals[category] = (ins, upd)

                print("\nIngestion summary:")
                for category, (ins, upd) in totals.items():
                    parsed, skipped = sheet_stats.get(category, (0, 0))
                    print(f"  {category}: parsed={parsed}, skipped={skipped}, inserted={ins}, updated={upd}")

    except Exception as exc:
        print(f"[ERROR] Database ingestion failed: {exc}", file=sys.stderr)
        return 1

    print("\nDone.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
