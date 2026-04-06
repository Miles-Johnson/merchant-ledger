#!/usr/bin/env python3
# -----------------------------------------------------------------------------
# DEPRECATED — superseded by parse_recipes_json.py as of 2026-04-04
# Retained for reference only. Do not run as part of the pipeline.
# -----------------------------------------------------------------------------

"""
Ingest raw recipe workbook rows into recipe_staging.

Environment variables:
  - DATABASE_URL: PostgreSQL connection string for psycopg2
  - RECIPES_XLS_PATH: Path to recipe workbook (.xlsx)
"""

import os
import sys
from typing import List, Sequence, Tuple

import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv


load_dotenv()


INSERT_SQL = """
INSERT INTO recipe_staging (
    source_mod,
    recipe_type,
    output_game_code,
    output_qty_raw,
    ingredients_raw
) VALUES (%s, %s, %s, %s, %s)
"""


def is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def parse_sheet_rows(sheet) -> List[Tuple[str, str, str, str, str]]:
    rows: List[Tuple[str, str, str, str, str]] = []

    # Skip row 1 (header):
    # Source Mod, Recipe Type, Output Item, Output Qty, Ingredients
    for values in sheet.iter_rows(min_row=2, values_only=True):
        source_mod = values[0] if len(values) > 0 else None
        recipe_type = values[1] if len(values) > 1 else None
        output_game_code = values[2] if len(values) > 2 else None
        output_qty_raw = values[3] if len(values) > 3 else None
        ingredients_raw = values[4] if len(values) > 4 else None

        # Skip any row where Output Item (column 2 / index 2) is blank
        if is_blank(output_game_code):
            continue

        rows.append(
            (
                as_text(source_mod),
                as_text(recipe_type),
                as_text(output_game_code),
                as_text(output_qty_raw),
                as_text(ingredients_raw),
            )
        )

    return rows


def main() -> int:
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        print("[ERROR] DATABASE_URL environment variable is not set.", file=sys.stderr)
        return 1

    recipes_xls_path = os.getenv("RECIPES_XLS_PATH")
    if not recipes_xls_path:
        print("[ERROR] RECIPES_XLS_PATH environment variable is not set.", file=sys.stderr)
        return 1

    print(f"Loading workbook: {recipes_xls_path}")
    try:
        workbook = load_workbook(filename=recipes_xls_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[ERROR] Failed to open workbook: {exc}", file=sys.stderr)
        return 1

    total_inserted = 0

    try:
        with psycopg2.connect(database_url) as conn:
            with conn.cursor() as cur:
                cur.execute("TRUNCATE TABLE recipe_staging")

                for sheet_name in workbook.sheetnames:
                    if sheet_name == "All Recipes":
                        print(f"Sheet '{sheet_name}': skipped (combined view)")
                        continue

                    sheet = workbook[sheet_name]
                    rows = parse_sheet_rows(sheet)

                    if rows:
                        cur.executemany(INSERT_SQL, rows)

                    inserted_count = len(rows)
                    total_inserted += inserted_count
                    print(f"Sheet '{sheet_name}': inserted {inserted_count} rows")

    except Exception as exc:
        print(f"[ERROR] Database ingestion failed: {exc}", file=sys.stderr)
        return 1
    finally:
        workbook.close()

    print(f"Total inserted: {total_inserted} rows")
    print("[OK] Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())